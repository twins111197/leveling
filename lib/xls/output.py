from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from functools import reduce
from operator import add

def output_master_excel(assignments, activities):
    """Output master excel document"""
    # Create the workbook
    book = Workbook()
    # Access the active sheet
    sheet = book.active
    sheet.title = "Bechirot"

    # Sets column widths
    columns = [1, 4]
    for column in columns:
        i = get_column_letter(column)
        sheet.column_dimensions[i].width = 23

    # Determine maximum number of past activities
    max_past_activities = max(len(camper.past_activities) for camper in assignments)

    # Create explicit header row -- Note: rows and columns indexed starting from 1 in openpyxl
    header = ["Name", "Edah", "Tzrif", "Peulah", "Preference"]
    if max_past_activities != 0:
        # Reduce takes an iterator and an operator and applies the operator one by one
        header += reduce(add, (["Past Peulah %d" % i, "Past Preference %d" % i] for i in range(1, max_past_activities + 1)))

    # Write header row
    for i, item in enumerate(header):
        sheet.cell(row=1, column=i+1).value = item


    # Create a list representing the row
    for row, camper in enumerate(assignments):
        attributes = [(camper.name, None, None),
                      (camper.edah, None, None),
                      (camper.bunk, None, None),
                      format_activity(camper, assignments[camper]),
                      (camper.pref_of(assignments[camper])+1, None, None)
                      ]

        # Add camper histories as attributes of a camper to be written
        if camper.past_activities != []:
            attributes += sum(
                ([ (activity.name, None, None), (preference, None, None) ]
                for activity, preference
                in zip(camper.past_activities, camper.past_preferences)),
                [])

        # This does the actual writing of information from the row list
        for col, (text, color, font) in enumerate(attributes):
            cell = sheet.cell(row=row + 2, column=col + 1)
            cell.value = text
            if color is not None:
                cell.fill = color
            if font is not None:
                cell.font = font



    # Write the second sheet of the output
    # Create 2nd worksheet
    sheet2 = book.create_sheet("Output", 0)

    # Sets column width for first 2 columns
    columns = [1, 2, 3, 4, 5, 6, 7, 8]
    for column in columns:
        i = get_column_letter(column)
        sheet2.column_dimensions[i].width = 23

    # Write header objects
    header = ["Peulah", "Name"]
    header += ["Preference %d" % i for i in range(1,7)]

    # Write header row
    for i, item in enumerate(header):
        sheet2.cell(row=1, column=i+1).value = item

    # Create groups and determine empty spots
    groups = group_by_activity(assignments, activities)

    # Write information to the sheet
    row = 2
    for activity, group in groups.items():
        for camper in group:
            if activity is not None:
                sheet2.cell(row=row, column=1).value = activity.name
            else:
                sheet2.cell(row=row, column=1).value = 'Unassigned'
            camper_cell = sheet2.cell(row=row, column=2)
            if camper is None:
                camper_cell.fill = PatternFill(start_color='2BFFF5', end_color='2BFFF5', fill_type = "solid")
            else:
                camper_cell.value = camper.name
                if activity is None:
                    camper_cell.fill = PatternFill(start_color='808080', end_color='808080', fill_type = "solid")
                    camper_cell.font = Font(color="FFFFFF")
                elif activity not in camper.preferences:
                    camper_cell.fill = PatternFill(start_color='000000', end_color='000000', fill_type = "solid")
                    camper_cell.font = Font(color="FFFFFF")
                else:
                    camper_cell.fill = determine_color(camper.pref_of(activity) + 1)
                for column, preference in enumerate(camper.preferences):
                    sheet2.cell(row=row, column=column+3).value = preference.name
            row += 1
        row += 1

    return book



def group_by_activity(assignments, activities):
    """Takes a list of campers, returns dictionary with activities as keys, elements are a list of camper objects in that activity"""
    groups = { activity: [camper for camper in assignments if assignments[camper] == activity] for activity in activities + [None] }

    # Add empty slots
    for activity in groups:
        num_campers = len(groups[activity])
        if activity is not None and num_campers < activity.capacity:
            groups[activity] += [None] * (activity.capacity - num_campers)

    return groups

def format_activity(camper, activity):
    """Assign color to a cell in an Excel Document based on Camper preference"""
    # Create an alert color if a camper wasn't assigned an activity
    if activity is None:
        return 'Unassigned', PatternFill(start_color='808080', end_color='808080', fill_type = "solid"), Font(color="FFFFFF")
    elif activity not in camper.preferences:
        return activity.name, PatternFill(start_color='000000', end_color='000000', fill_type = "solid"), Font(color="FFFFFF")
    else:
        # We know that there's at least one element in campers.past_preferences
        return activity.name, determine_color(camper.pref_of(activity)+1), None



def determine_color(number):
    if number == 1:
        return PatternFill(start_color='0fc70f', end_color='0fc70f', fill_type = "solid")
    elif number == 2:
        return PatternFill(start_color='FFEE08', end_color='FFEE08', fill_type = "solid")
    elif number == 3:
        return PatternFill(start_color='dd9a1f', end_color='dd9a1f', fill_type = "solid")
    else:
        return PatternFill(start_color='ff0000', end_color='ff0000', fill_type = "solid")
