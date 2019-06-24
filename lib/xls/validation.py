# Standard Library
from collections import Counter

# Packages
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def check_for_input_errors(sheet, activ_sheet):
    """Takes in the preferences spreadsheet as sheet, activities sheet as activ_sheet, if there are errors then write them to an excel doc."""
    errors = []

    # Check that all names are unique
    names_ive_seen = set()  # curly braces is an empty set, so is set()
    for row in sheet.rows:
        if not row:
            continue
        name = row[0].value
        if name in names_ive_seen:
            errors.append(
                "At least two campers have the name %s in the preferences document"
                % name
            )
        else:
            names_ive_seen.add(name)

    # Check that no cells are empty
    for row in sheet.rows:
        for cell in row:
            if cell.value is None:
                errors.append("%s is empty." % cell.coordinate)

    # Check that all preferences are in the activities sheet
    activs = [cell.value for cell in activ_sheet['A']]
    print(activs)
    for row_cells in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for col_cells in sheet.iter_cols(min_col=4, max_col=sheet.max_column):
            for cell in col_cells:
                if cell in row_cells and cell in col_cells and cell.value not in activs:
                    errors.append("%s was a chanich's preference but is not an activity in the Activities Sheet." % cell.value)

    return errors


def output_errors(errors_list):
    """Takes in a list of errors, outputs an Excel document with these errors. Writes with Openpyxl."""
    # Create the workbook
    book = Workbook()
    # Access the active sheet
    sheet = book.active
    sheet.title = "Errors"
    for i, error in enumerate(errors_list):
        sheet.cell(row=i + 1, column=1).value = error

    return book
