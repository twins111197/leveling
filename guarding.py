import xlrd
from openpyxl import Workbook

def check_preferences_for_input_errors(sheet):
    """Takes in the preferences spreadsheet, if there are errors then write them to an excel doc. Reads with xlrd."""
    errors = []

    # Check that all names are unique
    for row1 in range(sheet.nrows):
        for row2 in range(sheet.nrows):
            if row1 != row2:
                if sheet.cell(row1, 0).value == sheet.cell(row2, 0).value:
                    if "At least two campers have the name %s in the preferences document" % sheet.cell(row1, 0).value not in errors:
                        errors.append("At least two campers have the name %s in the preferences document" % sheet.cell(row1, 0).value)


    return errors



def output_errors(errors_list):
    """Takes in a list of errors, outputs an Excel document with these errors. Writes with Openpyxl."""
    # Create the workbook
    book = Workbook()
    # Access the active sheet
    sheet = book.active
    sheet.title = "Errors"
    for i, error in enumerate(errors_list):
        sheet.cell(row=i+1, column=1).value = error

    return book
